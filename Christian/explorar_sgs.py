import pandas as pd
from openpyxl import load_workbook
import warnings
warnings.filterwarnings("ignore")

ARCHIVO_SEGUIMIENTO = "PRUEBAS/Seguimiento PROGRAMAS FTTH-TBA_27ABR26.xlsm"

# ===================== Explorar documentos externos (son TSV disfrazados de .xls) =====================
print("=" * 80)
print("DOCUMENTO EXTERNO: SGS-PON NUEVO.xls (TSV)")
print("=" * 80)
df_nuevo = pd.read_csv("PRUEBAS/SGS-PON NUEVO.xls", sep="\t", encoding="latin-1")
print(f"Shape: {df_nuevo.shape}")
print(f"Columnas ({len(df_nuevo.columns)}):")
for i, col in enumerate(df_nuevo.columns):
    print(f"  {i}: '{col}'")
print(f"\nPrimeras 2 filas:")
print(df_nuevo.head(2).to_string())

print("\n" + "=" * 80)
print("DOCUMENTO EXTERNO: SGS-PON EXISTENTE.xls (TSV)")
print("=" * 80)
df_existente = pd.read_csv("PRUEBAS/SGS-PON EXISTENTE.xls", sep="\t", encoding="latin-1")
print(f"Shape: {df_existente.shape}")
print(f"Columnas ({len(df_existente.columns)}):")
for i, col in enumerate(df_existente.columns):
    print(f"  {i}: '{col}'")
print(f"\nPrimeras 2 filas:")
print(df_existente.head(2).to_string())

# ===================== Explorar pestanas del Seguimiento =====================
wb = load_workbook(ARCHIVO_SEGUIMIENTO, read_only=True, keep_vba=True)

for sheet_name in ["SGS-PON NUEVO", "SGS-PON EXISTENTE"]:
    print("\n" + "=" * 80)
    print(f"PESTAÑA EN SEGUIMIENTO: '{sheet_name}'")
    print("=" * 80)
    ws = wb[sheet_name]
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    
    # Leer fila 1 (encabezados)
    print(f"\nFila 1 (encabezados):")
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val is not None:
            print(f"  Col {col}: '{val}'")
    
    # Leer primeras filas de datos
    for row_n in [2, 3]:
        print(f"\nFila {row_n}:")
        for col in range(1, min(ws.max_column + 1, 25)):
            val = ws.cell(row=row_n, column=col).value
            if val is not None:
                print(f"  Col {col}: '{val}'")

    # Ver si hay tablas
    # Necesitamos abrir sin read_only para ver tablas
    
wb.close()

# Re-abrir para ver tablas
wb2 = load_workbook(ARCHIVO_SEGUIMIENTO, keep_vba=True, data_only=True)
for sheet_name in ["SGS-PON NUEVO", "SGS-PON EXISTENTE"]:
    ws2 = wb2[sheet_name]
    print(f"\nTablas en '{sheet_name}': {list(ws2.tables.keys()) if ws2.tables else 'Ninguna'}")
    for t in ws2.tables.values():
        print(f"  Tabla '{t.displayName}': ref={t.ref}")
wb2.close()

# ===================== Comparar columnas =====================
print("\n" + "=" * 80)
print("COMPARACION DE COLUMNAS")
print("=" * 80)

# Leer las pestanas como dataframes
df_tab_nuevo = pd.read_excel(ARCHIVO_SEGUIMIENTO, sheet_name="SGS-PON NUEVO")
df_tab_existente = pd.read_excel(ARCHIVO_SEGUIMIENTO, sheet_name="SGS-PON EXISTENTE")

print(f"\nPestaña SGS-PON NUEVO - shape: {df_tab_nuevo.shape}")
print(f"Columnas ({len(df_tab_nuevo.columns)}):")
for i, col in enumerate(df_tab_nuevo.columns):
    print(f"  {i}: '{col}'")

print(f"\nPestaña SGS-PON EXISTENTE - shape: {df_tab_existente.shape}")
print(f"Columnas ({len(df_tab_existente.columns)}):")
for i, col in enumerate(df_tab_existente.columns):
    print(f"  {i}: '{col}'")

# Columnas en comun
cols_nuevo_doc = set(str(c).strip() for c in df_nuevo.columns)
cols_nuevo_tab = set(str(c).strip() for c in df_tab_nuevo.columns)
print(f"\n--- SGS-PON NUEVO ---")
print(f"Columnas en AMBOS: {sorted(cols_nuevo_doc & cols_nuevo_tab)}")
print(f"Solo en documento: {sorted(cols_nuevo_doc - cols_nuevo_tab)}")
print(f"Solo en pestaña: {sorted(cols_nuevo_tab - cols_nuevo_doc)}")

cols_existente_doc = set(str(c).strip() for c in df_existente.columns)
cols_existente_tab = set(str(c).strip() for c in df_tab_existente.columns)
print(f"\n--- SGS-PON EXISTENTE ---")
print(f"Columnas en AMBOS: {sorted(cols_existente_doc & cols_existente_tab)}")
print(f"Solo en documento: {sorted(cols_existente_doc - cols_existente_tab)}")
print(f"Solo en pestaña: {sorted(cols_existente_tab - cols_existente_doc)}")
