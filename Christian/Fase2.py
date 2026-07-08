import pandas as pd
from datetime import date, timedelta
import numpy as np
from openpyxl import load_workbook  # CAMBIO: Importar openpyxl para editar el .xlsm
from openpyxl.utils import get_column_letter  # Para actualizar rangos de tablas

pd.set_option('display.max_rows', None)

# ===================== Nombre del archivo Seguimiento (centralizado) =====================
# CAMBIO: Variable centralizada para el nombre del archivo, asi se usa en pandas Y openpyxl
ARCHIVO_SEGUIMIENTO = "PRUEBAS/Seguimiento PROGRAMAS FTTH-TBA_27ABR26.xlsm"
ARCHIVO_SALIDA = "PRUEBAS/Seguimiento_ACTUALIZADO.xlsm"  # CAMBIO: Nombre del archivo de salida


#Tabla con Nuevos Datos a agregar 
df_ND = pd.read_excel("PRUEBAS/Distritos HOY.XLSX")
df_ND["SGS"] = df_ND["SGS"].fillna('N/A', inplace=False)
df_ND = df_ND.iloc[:, :6]
for i in range(len(df_ND)):

    if "." in str(df_ND.loc[i, "Terminal"]):
        df_ND.loc[i, "Terminal"] = df_ND.loc[i, "Terminal"].replace(".", ",")




Sem = date.today().isocalendar().week
M =  14
print("Semana Actual:")
print(Sem)

#Tabla Principal
df_TP = pd.read_excel(ARCHIVO_SEGUIMIENTO, sheet_name="BD")  # CAMBIO: Usar variable centralizada
df_TP = df_TP.iloc[12:].reset_index(drop=True)
df_TP = df_TP.drop(df_TP.columns[:1], axis=1)
df_TP.columns = df_TP.iloc[0]
df_TP = df_TP.iloc[1:].reset_index(drop=True)
df_TP = df_TP.iloc[:, :19]
df_TP = df_TP[df_TP["ETAPA SGS ACTUALIZADO"] != "En Servicio"]
df_TP = df_TP[["SEM PROG", "DISTRITO", "SGS", "Terminales", "ETAPA SGS ACTUALIZADO", "PUENTES"]]
df_TP = df_TP[df_TP['SEM PROG'].isin([Sem])].reset_index(drop=True)
df_TP["PUENTES"] = df_TP["PUENTES"].fillna('Nuevo', inplace=False)
for i in range(len(df_TP)):

    if "." in df_TP.loc[i, "Terminales"]:
        df_TP.loc[i, "Terminales"] = df_TP.loc[i, "Terminales"].replace(".", ",")



y = 0
df_C = pd.DataFrame(columns=["BUSCA X SGS","B X DISTRITO", "Ter x Dis", "¿Igual?"])

# Busqueda de SGS
for i in range(len(df_ND)):
    for n in range(len(df_TP)):
        if df_ND.loc[i, "SGS"] == "N/A":
            df_C.loc[y, "BUSCA X SGS"] = "N/A"
            y += 1
            break
        elif df_ND.loc[i, "SGS"] == df_TP.loc[n, "SGS"]:
            df_C.loc[y, "BUSCA X SGS"] = df_TP.loc[n, "Terminales"]
            y += 1
            break
    else:
        y += 1
        
y = 0

# Busqueda de Distrito
for i in range(len(df_ND)):
    P = None
    for n in range(len(df_TP)):
        
        if df_ND.loc[i, "DIST"] == df_TP.loc[n, "DISTRITO"]:
            if P == None:
                P = df_TP.loc[n, "SGS"]
                if P == df_ND.loc[i, "SGS"]:
                    break 


            else:
                P = df_TP.loc[n, "SGS"]
                if P == df_ND.loc[i, "SGS"]:
                    break 

    if P == None:
        df_C.loc[y, "B X DISTRITO"] = "N/A"
        y += 1
    else:
        df_C.loc[y, "B X DISTRITO"] = P
        y += 1

y = 0
LD = pd.DataFrame(columns=["SGS", "Distrito", "Terminales"])


# Busqueda de Terminales por Distrito
for i in range(len(df_ND)):
    P = None
    for n in range(len(df_TP)):
        if df_ND.loc[i, "DIST"] == df_TP.loc[n, "DISTRITO"]:
            if P == None:
                P = df_TP.loc[n, "Terminales"]
                if P[:2] in df_ND.loc[i, "Terminal"]:
                    break

                else:
                    LD.loc[len(LD)] = [df_TP.loc[n, "SGS"], df_TP.loc[n, "DISTRITO"], df_TP.loc[n, "Terminales"]]

            else:
                P = df_TP.loc[n, "Terminales"]
                if P[:2] in df_ND.loc[i, "Terminal"]:
                    break
                else:
                    LD.loc[len(LD)] = [df_TP.loc[n, "SGS"], df_TP.loc[n, "DISTRITO"], df_TP.loc[n, "Terminales"]]



    if P == None:
        df_C.loc[y, "Ter x Dis"] = "N/A"
        y += 1

    else:        
        df_C.loc[y, "Ter x Dis"] = P
        y += 1



LD.drop_duplicates(inplace=True)
LD.reset_index(drop=True, inplace=True)
print(LD)

df_C["BUSCA X SGS"] = df_C["BUSCA X SGS"].fillna('N/A', inplace=False)
df_C["B X DISTRITO"] = df_C["B X DISTRITO"].fillna('N/A', inplace=False)

# Comparacion
for i in range(len(df_C)):
    if df_C.loc[i, "BUSCA X SGS"] == "N/A" and df_C.loc[i, "B X DISTRITO"] == "N/A" and df_C.loc[i, "Ter x Dis"] == "N/A":
        df_C.loc[i, "¿Igual?"] = "N/A"
    elif df_C.loc[i, "B X DISTRITO"] == df_ND.loc[i, "SGS"] and (df_C.loc[i, "Ter x Dis"] == df_ND.loc[i, "Terminal"] or df_C.loc[i, "Ter x Dis"] == df_C.loc[i, "BUSCA X SGS"]):
        df_C.loc[i, "¿Igual?"] = "SI"
    else:
        df_C.loc[i, "¿Igual?"] = "NO"

df = pd.concat([df_ND, df_C], axis=1)

print("Tabla Completa Antes de la Clasificacion:\n")
print(df,"\n")
print("--"*50,"\n")

df = df[df['¿Igual?'].isin(["NO", "N/A"])].reset_index(drop=True)

print("Tabla Completa Despues de la Clasificacion:\n")
print(df,"\n")
print("--"*50,"\n")

i = 0
while i < len(df):
    if df.loc[i, "Terminal"] == df.loc[i, "Ter x Dis"] and df.loc[i, "SGS"] != df.loc[i, "B X DISTRITO"] and df.loc[i, "SGS"] == "N/A":
        df.drop(i, inplace=True)
        df.reset_index(drop=True, inplace=True)
    elif df.loc[i, "SGS"] != df.loc[i, "B X DISTRITO"] and df.loc[i, "SGS"] == "N/A" and df.loc[i, "Terminal"]  in df.loc[i, "Ter x Dis"]: 
        df.drop(i, inplace=True)
        df.reset_index(drop=True, inplace=True)
    else:
        i += 1

df["B X DISTRITO"] = df["B X DISTRITO"].fillna('N/A', inplace=False)

df_SGS = pd.DataFrame(columns=["SGS", "DIST", "Pon", "Terminal", "ETAPA SGS ACTUALIZADO", "Comentarios"])

# Solo falta SGS

for i in range(len(df)):
    if df.loc[i, "B X DISTRITO"] == "N/A" and df.loc[i, "SGS"] != "N/A":
        if df.loc[i, "Terminal"] == df.loc[i, "Ter x Dis"]:
            df_SGS.loc[len(df_SGS)] = [df.loc[i, "SGS"], df.loc[i, "DIST"], df.loc[i, "Pon"], df.loc[i, "Terminal"], "", ""]
    
        else:
            J = False
            for n in range(len(LD)):
                if df.loc[i, "DIST"] == LD.loc[n, "Distrito"]:

                    if LD.loc[n, "Terminales"][:2] in df.loc[i, "Terminal"]:
                        J = True
                        break


            
            if J:
                df_SGS.loc[len(df_SGS)] = [df.loc[i, "SGS"], df.loc[i, "DIST"], df.loc[i, "Pon"], df.loc[i, "Terminal"], "", "Validar terminales"]

        
           

#Eliminar registros ya clasificados

for i in (df_SGS.index):
    for n in (df).index:
        if df_SGS.loc[i, "SGS"] == df.loc[n, "SGS"] and df_SGS.loc[i, "DIST"] == df.loc[n, "DIST"] and df_SGS.loc[i, "Terminal"] == df.loc[n, "Terminal"]:
            df.drop(n, inplace=True)
            df.reset_index(drop=True, inplace=True)
            break

# Nuevos Registros

df_NR = pd.DataFrame(columns=["PROG", "SGS", "DIST", "PTOS", "Pon", "Terminal", "ETAPA SGS ACTUALIZADO", "Comentarios"])


for i in range(len(df)):
    if df.loc[i, "¿Igual?"] == "N/A":
        df_NR.loc[len(df_NR)] = [df.loc[i, "PROG"], df.loc[i, "SGS"], df.loc[i, "DIST"], df.loc[i, "PTOS"], df.loc[i, "Pon"], df.loc[i, "Terminal"], "", ""]


    elif df.loc[i, "Terminal"] != df.loc[i, "Ter x Dis"]:
        df_NR.loc[len(df_NR)] = [df.loc[i, "PROG"], df.loc[i, "SGS"], df.loc[i, "DIST"], df.loc[i, "PTOS"], df.loc[i, "Pon"], df.loc[i, "Terminal"], "", ""]
    elif df.loc[i, "B X DISTRITO"] != "N/A" and df.loc[i, "B X DISTRITO"] != df.loc[i, "SGS"] and df.loc[i, "SGS"] != "N/A":
        df_NR.loc[len(df_NR)] = [df.loc[i, "PROG"], df.loc[i, "SGS"], df.loc[i, "DIST"], df.loc[i, "PTOS"], df.loc[i, "Pon"], df.loc[i, "Terminal"], "", ""]

# Eliminar registros ya clasificados

for i in range(len(df_NR)):
    for n in range(len(df)):
        if df_NR.loc[i, "PROG"] == df.loc[n, "PROG"] and df_NR.loc[i, "SGS"] == df.loc[n, "SGS"] and df_NR.loc[i, "DIST"] == df.loc[n, "DIST"] and df_NR.loc[i, "Terminal"] == df.loc[n, "Terminal"]:
            df.drop(n, inplace=True)
            df.reset_index(drop=True, inplace=True)
            break

df_Prioridad = pd.read_excel("PRUEBAS/CATALOGO DISTRITOS PRIORITARIOS SEM15 2026.xlsx")

for i in range(len(df_NR)):
    for n in range(len(df_Prioridad)):
        if df_NR.loc[i, "DIST"] == df_Prioridad.loc[n, "Distrito"]:
            df_NR.loc[i, "Comentarios"] = "Prioridad"
            break

for i in range(len(df_SGS)):
    for n in range(len(df_Prioridad)):
        if df_SGS.loc[i, "DIST"] == df_Prioridad.loc[n, "Distrito"]:
            if df_SGS.loc[i, "Comentarios"] == "":
                df_SGS.loc[i, "Comentarios"] = "Prioridad"

            else:
                df_SGS.loc[i, "Comentarios"] = "Prioridad, " + df_SGS.loc[i, "Comentarios"]

            break




print("Faltantes por clasificar:\n")
print(df,"\n")
print("--"*50,"\n")            

print("Registrar solamente la clave SGS:\n")
print(df_SGS,"\n")
print("--"*50,"\n")

print("Nuevos Registros:\n")
print(df_NR,"\n")
print("--"*50,"\n")


# ================================================== FASE 2 ==================================================

print("================================================== FASE 2 ==================================================")

# Abrir el workbook con openpyxl preservando macros (keep_vba=True)
wb = load_workbook(ARCHIVO_SEGUIMIENTO, keep_vba=True)
ws = wb["BD"]

# Leer la fila de encabezados (fila 14) para mapear nombres de columna a indices
col_map = {}  # Diccionario: nombre_columna -> numero_columna (1-indexed)
for col in range(1, ws.max_column + 1):
    valor = ws.cell(row=14, column=col).value
    if valor != None:
        col_map[str(valor).strip()] = col

print("Columnas encontradas en fila 14:")
for nombre, idx in col_map.items():
    print(f"  Col {idx}: '{nombre}'")

# ESTRATEGIA: Insertar filas nuevas al INICIO del area de datos (fila 15)
# y recorrer toda la informacion existente hacia abajo.
# Esto evita tener que detectar la ultima fila con datos.

FILA_INICIO_DATOS = 15  # Primera fila de datos (despues de encabezados en fila 14)
cantidad_nuevas = len(df_NR)

if cantidad_nuevas > 0:
    print(f"\nInsertando {cantidad_nuevas} filas nuevas en la fila {FILA_INICIO_DATOS} (los datos existentes se recorren hacia abajo)...")
    
    # Insertar filas vacias en la fila 15, esto empuja todo lo existente hacia abajo
    ws.insert_rows(FILA_INICIO_DATOS, amount=cantidad_nuevas)
    
    # Actualizar el rango de todas las tablas en la hoja para incluir las filas insertadas
    # insert_rows NO actualiza automaticamente los objetos Table de Excel,
    # lo que causa que las filas recorridas queden fuera de la tabla y no se filtren.
    for tabla in ws.tables.values():
        ref_original = tabla.ref  # Ejemplo: "B14:S5382"
        partes = ref_original.split(":")
        celda_inicio = partes[0]  # "B14"
        celda_fin = partes[1]      # "S5382"
        
        # Extraer la fila final del rango
        col_fin_letras = ''.join(c for c in celda_fin if c.isalpha())
        fila_fin = int(''.join(c for c in celda_fin if c.isdigit()))
        
        # Sumar las filas insertadas al rango final
        nueva_fila_fin = fila_fin + cantidad_nuevas
        nuevo_ref = f"{celda_inicio}:{col_fin_letras}{nueva_fila_fin}"
        
        print(f"  Tabla '{tabla.displayName}': {ref_original} -> {nuevo_ref}")
        tabla.ref = nuevo_ref
    
    # Escribir los datos de df_NR en las filas recien insertadas
    for i in range(cantidad_nuevas):
        fila_actual = FILA_INICIO_DATOS + i
        
        ws.cell(row=fila_actual, column=col_map["GENERICO"], value="7. OTROS")

        ws.cell(row=fila_actual, column=col_map["PROGRAMA"], value=df_NR.loc[i, "PROG"])

        ws.cell(row=fila_actual, column=col_map["Prioridad etapa"], value="=VLOOKUP([@[ETAPA SGS ACTUALIZADO]],cat_eFTTH!F:G,2,0)")

        ws.cell(row=fila_actual, column=col_map["SEM PROG"], value=Sem)

        # Escribir DISTRITO 
        ws.cell(row=fila_actual, column=col_map["DISTRITO"], value=df_NR.loc[i, "DIST"])
        
        # Escribir SGS
        ws.cell(row=fila_actual, column=col_map["SGS"], value=df_NR.loc[i, "SGS"])

        if df_NR.loc[i, "Comentarios"] == "Prioridad":
            ws.cell(row=fila_actual, column=col_map["PRIORIDAD"], value="1")

        
        # Escribir Terminales 
        ws.cell(row=fila_actual, column=col_map["Terminales"], value=df_NR.loc[i, "Terminal"])

        ws.cell(row=fila_actual, column=col_map["Puertos Construidos"], value=df_NR.loc[i, "PTOS"])

        ws.cell(row=fila_actual, column=col_map["ZONA COMERCIAL"] , value="=VLOOKUP([@DISTRITO],GZ!A:F,6,0)")

        ws.cell(row=fila_actual, column=col_map["COPE"] , value="=VLOOKUP([@DISTRITO],GZ!A:I,9,0)")

        ws.cell(row=fila_actual, column=col_map["LOCALIDAD"] , value="=VLOOKUP([@DISTRITO],GZ!A:K,11,0)")

        if df_NR.loc[i, "Pon"] == "Nuevo":
            ws.cell(row=fila_actual, column=col_map["ETAPA SGS ACTUALIZADO"] , value='=IFERROR(VLOOKUP([@SGS],"SGS-PON NUEVO"!B:U,20,0),"PENDIENTE SGS")')
        else:
            ws.cell(row=fila_actual, column=col_map["ETAPA SGS ACTUALIZADO"] , value='=IFERROR(VLOOKUP([@SGS],"SGS-PON EXISTENTE"!B:U,20,0),"PENDIENTE SGS")')

        ws.cell(row=fila_actual, column=col_map["PUENTES"] , value=df_NR.loc[i, "Pon"])

        
        print(f"  Fila {fila_actual}: DIST={df_NR.loc[i, 'DIST']}, SGS={df_NR.loc[i, 'SGS']}, Terminal={df_NR.loc[i, 'Terminal']}")
else:
    print("\nNo hay filas nuevas que insertar.")

print(f"\nTotal de filas insertadas en BD: {cantidad_nuevas}")
print("================================================== FIN FASE 2 ==================================================")

# ================================================== FASE 3 ==================================================

print("\n================================================== FASE 3 ==================================================")
print("Actualizando pestañas SGS-PON NUEVO y SGS-PON EXISTENTE...")

# Leer documentos externos (son TSV disfrazados de .xls)
df_sgs_nuevo = pd.read_csv("PRUEBAS/SGS-PON NUEVO.xls", sep="\t", encoding="latin-1")
df_sgs_existente = pd.read_csv("PRUEBAS/SGS-PON EXISTENTE.xls", sep="\t", encoding="latin-1")

# Eliminar columna vacia al final si existe
for df_temp in [df_sgs_nuevo, df_sgs_existente]:
    cols_unnamed = [c for c in df_temp.columns if "Unnamed" in str(c)]
    if cols_unnamed:
        df_temp.drop(columns=cols_unnamed, inplace=True)

# Configuracion por pestaña
config_pestanas = [
    {"nombre": "SGS-PON NUEVO", "df": df_sgs_nuevo, "tabla": "Table2"},
    {"nombre": "SGS-PON EXISTENTE", "df": df_sgs_existente, "tabla": "Table3"},
]

for config in config_pestanas:
    sheet_name = config["nombre"]
    df_datos = config["df"]
    tabla_nombre = config["tabla"]
    
    print(f"\n--- Procesando pestaña: '{sheet_name}' ---")
    ws_sgs = wb[sheet_name]
    
    # Mapear encabezados de la fila 1 a indices de columna
    col_map_sgs = {}
    for col in range(1, ws_sgs.max_column + 1):
        val = ws_sgs.cell(row=1, column=col).value
        if val is not None:
            col_map_sgs[str(val).strip()] = col
    
    # Identificar las columnas comunes (las 19 que vienen en el documento)
    columnas_comunes = [c for c in df_datos.columns if c in col_map_sgs]
    print(f"  Columnas comunes a sobreescribir: {len(columnas_comunes)}")
    
    filas_actuales = ws_sgs.max_row - 1  # Restar header
    filas_nuevas = len(df_datos)
    print(f"  Filas actuales: {filas_actuales}, Filas nuevas: {filas_nuevas}")
    

    for i in range(filas_nuevas):
        fila_excel = i + 2  # Fila 2 en adelante (fila 1 es header)
        for col_nombre in columnas_comunes:
            col_idx = col_map_sgs[col_nombre]
            valor = df_datos.iloc[i][col_nombre]
            # Convertir NaN a None
            if pd.isna(valor):
                valor = None
            ws_sgs.cell(row=fila_excel, column=col_idx, value=valor)
    
    # Actualizar rango de la tabla
    for tabla in ws_sgs.tables.values():
        if tabla.displayName == tabla_nombre:
            ref_original = tabla.ref
            partes = ref_original.split(":")
            celda_inicio = partes[0]
            celda_fin = partes[1]
            col_fin_letras = ''.join(c for c in celda_fin if c.isalpha())
            nueva_fila_fin = filas_nuevas + 1  # +1 por header
            nuevo_ref = f"{celda_inicio}:{col_fin_letras}{nueva_fila_fin}"
            print(f"  Tabla '{tabla_nombre}': {ref_original} -> {nuevo_ref}")
            tabla.ref = nuevo_ref
    
    print(f"  Completado: {filas_nuevas} filas escritas en '{sheet_name}'")

print("\n================================================== FIN FASE 3 ==================================================")

# Guardar el archivo como copia (NO sobreescribir el original)
wb.save(ARCHIVO_SALIDA)

print(f"\nArchivo guardado exitosamente como: {ARCHIVO_SALIDA}")
print("================================================== FIN ==================================================")
